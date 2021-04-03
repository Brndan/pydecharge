#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Tous les fichiers, en entrée comme en sortie, sont des XLSX.


import sys
import os
from pathlib import Path
import shutil
import argparse
import re
import glob

from openpyxl import Workbook, load_workbook


def check_row(row, file, line):
    line += 1
    champs = dict(zip(["civilite", "prenom", "nom", "heures_decharge",
                       "minutes_decharges", "heures_ORS", "corps", "rne"], row))
    if not champs['civilite'] in ('M.', 'Mme'):
        print("Fichier {} : erreur dans le champ Civilité de la ligne {} traitée.".format(
            file, line), file=sys.stderr)
    if not re.fullmatch(r"""[A-ZÀ-ÖØ-ÞŽŸa-zà-öø-ÿ '-]+[a-zà-öø-ÿ]""", champs['prenom']):
        print("Fichier {} : erreur dans le champ Prénom de la ligne {} traitée.".format(
            file, line), file=sys.stderr)
    if not re.fullmatch(r"""[A-ZÀ-ÖØ-ÞŽŸ' -]+[A-ZÀ-ÖØ-ÞŽŸ]""", champs['nom']):
        print("Fichier {} : erreur dans le champ Nom de la ligne {} traitée.".format(
            file, line), file=sys.stderr)
    if not re.fullmatch(r'\d{2,3}', champs['corps']):
        print("Fichier {} : erreur dans le champ Corps de la ligne {} traitée.".format(
            file, line), file=sys.stderr)
    if not re.fullmatch(r'\d{7}[A-Z]', champs['rne']):
        print("Fichier {} : erreur dans le champ RNE de la ligne {} traitée.".format(
            file, line), file=sys.stderr)
    if champs['minutes_decharges'] < 0 or champs['minutes_decharges'] >= 60:
        print("Fichier {} : erreur dans le champ Minutes de décharge de la ligne {} traitée.".format(
            file, line), file=sys.stderr)
    if not champs['heures_ORS'] in (15, 18, 27, 35, 36, 1607):
        print("Fichier {} : erreur dans le champ Heures ORS de la ligne {} traitée.".format(
            file, line), file=sys.stderr)


def save_export_syndicats(export_sheet, output_file_path):
    wb = Workbook()
    ws = wb.active
    header = ["Code organisation", "M. Mme", "Prénom", "Nom", "Heures décharges",
              "Minutes décharges", "Heures ORS", "Minutes ORS", "AIRE", "Corps", "RNE"]
    for i in range(len(header)):
        ws.cell(1, i+1).value = header[i]
    # On remplit le fichier ici avec le contenu de export_sheet
    for row in range(len(export_sheet)):
        export_sheet[row].insert(0, "S01")  # Code organisation → toujours S01
        export_sheet[row].insert(7, 0)  # Minutes ORS → toujours 0
        export_sheet[row].insert(-2, 2)  # Aire, toujours 2
        for cell in range(len(export_sheet[row])):
            ws.cell(row+2, cell+1).value = export_sheet[row][cell]
            # Appliquer le format 'Texte' évite une single quote avant les nombres
            ws.cell(row+2, cell+1).number_format = '@'
    wb.save(output_file_path)

def save_export_cts(export_sheet, output_file_path):
    wb = Workbook()
    ws = wb.active
    header = ["Syndicat", "ETP attribués", "Mutualisation", "ETP disponibles", "Décharges",
              "CTS"]
    for i in range(len(header)):
        ws.cell(1, i+1).value = header[i]
    # On remplit le fichier ici avec le contenu de export_sheet
    for row in range(len(export_sheet)):
        for cell in range(len(export_sheet[row])):
            ws.cell(row+2, cell+1).value = export_sheet[row][cell]
            # Appliquer le format 'Texte' évite une single quote avant les nombres
            #ws.cell(row+2, cell+1).number_format = '@'
    wb.save(output_file_path)


def main():
    parser = argparse.ArgumentParser(
        description="Produire à partir d’un modèle les fichiers de décharge pour tous les syndicats.")
    parser.add_argument(
        "--cts",
        help="Synthèse de la ligne des CTS des syndicats",
        action="store_true",
        required=False)
    parser.add_argument(
        "--begin",
        "-b",
        action="store",
        help="Début de la plage de données",
        required=True
    )
    parser.add_argument(
        "--end",
        "-e",
        action="store",
        help="Fin de la plage de données",
        required=True
    )
    parser.add_argument(
        "-i",
        "--input",
        action="store",
        help="dossier dans lequel se trouvent les fichiers à compiler. Si ce paramètre est omis, le répertoire courant est utilisé.",
        required=False
    )
    parser.add_argument(
        "-o",
        "--output",
        action="store",
        help="fichier en sortie. Par défaut, le fichier export.xlsx est généré dans le répertoire courant.",
        required=False
    )
    """ parser.add_argument(
        "--csv",
        help="Synthèse de la ligne des CTS des syndicats",
        required=False) """
    args = parser.parse_args()

    if args.input:
        source_folder = Path(args.input)
    else:
        source_folder = os.getcwd()

    if args.output:
        output_file_path = Path(args.output)
    else:
        output_file_path = os.path.join(os.getcwd(), "export.xlsx")

    if not args.begin:
        args.begin = "B25"
    if not args.end:
        args.end = "J44"

    if not os.path.exists(source_folder):
        sys.exit("La source spécifiée n’existe pas.")

    # Récupère une liste de tous les xlsx dans le dossier source
    all_xlsx = glob.glob(os.path.join(source_folder, "*.xlsx"))

    export_sheet = []
    if args.cts:
        for xlsx_file in all_xlsx:
            try:
                file = load_workbook(filename=xlsx_file,data_only=True) # Récupérer les valeurs calculées
            except:
                sys.exit("Erreur à l’ouverture du fichier " + xlsx_file)
            sheet = file.active
            cell_range = sheet[args.begin:args.end]
            for row in range(len(cell_range)):
                export_row = []
                row_length = len(cell_range[row])
                empty_cells = 0
                for cell_coordinate in range(row_length):
                    cell = cell_range[row][cell_coordinate].value
                    export_row.append(cell)
                export_sheet.append(export_row)
        save_export_cts(export_sheet, output_file_path)
    else:
        for xlsx_file in all_xlsx:
            try:
                file = load_workbook(filename=xlsx_file)
            except:
                sys.exit("Erreur à l’ouverture du fichier " + xlsx_file)
            sheet = file.active
            cell_range = sheet[args.begin:args.end]
            for row in range(len(cell_range)):
                export_row = []
                row_length = len(cell_range[row])
                empty_cells = 0
                for cell_coordinate in range(row_length):
                    cell = cell_range[row][cell_coordinate].value
                    if not cell:
                        cell = 0
                        empty_cells += 1
                    export_row.append(cell)
                if not empty_cells == row_length:
                    check_row(export_row, xlsx_file, row)
                    export_sheet.append(export_row)
        save_export_syndicats(export_sheet, output_file_path)


if __name__ == "__main__":
    main()
